"""Extract SCAPS reference sweep data from partner xlsx into JSON.

Source: docs/superpowers/references/scaps_1r_parameters.xlsx (12 sheets)
Output: perovskite-sim/tests/integration/scaps_reference.json

Base point anchored from scaps_1d_simulation_report.pdf page 2.

Run from repo root:
    cd perovskite-sim && python scripts/extract_scaps_reference.py
"""
from __future__ import annotations

import json
import math
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = REPO_ROOT / "docs" / "superpowers" / "references" / "scaps_1r_parameters.xlsx"
JSON_OUT = REPO_ROOT / "perovskite-sim" / "tests" / "integration" / "scaps_reference.json"

BASE_MODEL = {
    "Voc_V": 1.1676,
    "Jsc_mA_cm2": 26.281994,
    "FF_percent": 86.99,
    "PCE_percent": 26.69,
}

SCALAR_SWEEP_SHEETS = {
    "CHI_ETL": {"x_col": "delta_E_C_eV", "x_unit": "eV"},
    "Et_HTL PVK": {"x_col": "scan_value", "x_unit": "eV"},
    "Et_C_PVK": {"x_col": "scan_value", "x_unit": "eV"},
    "Et_PVK ETL": {"x_col": "scan_value", "x_unit": "eV"},
    "Et_V_PVK": {"x_col": "scan_value", "x_unit": "eV"},
    "Nd_ETL": {"x_col": "scan_value", "x_unit": "cm^-3"},
    "Nt_HTL PVK": {"x_col": "scan_value", "x_unit": "cm^-3"},
    "Nt_C_PVK": {"x_col": "scan_value", "x_unit": "cm^-3"},
    "Nt_PVK ETL": {"x_col": "scan_value", "x_unit": "cm^-3"},
    "Nt_V_PVK": {"x_col": "scan_value", "x_unit": "cm^-3"},
}

PAIR_SHEETS = ("pair-nt-cbo-pvk-etl", "pair-nt-et-pvk-etl")

METRIC_COLS = ("PCE_percent", "Voc_V", "FF_percent", "Jsc_mA_cm2")


def _row_dict(ws, header: list[str], row_idx: int) -> dict:
    return {header[i]: ws.cell(row_idx, i + 1).value for i in range(len(header))}


def _clean_metrics(rd: dict) -> dict | None:
    out = {}
    for k in METRIC_COLS:
        v = rd.get(k)
        if v is None or (isinstance(v, str) and v.strip().upper() == "NA"):
            return None
        if not isinstance(v, (int, float)) or (isinstance(v, float) and math.isnan(v)):
            return None
        out[k] = float(v)
    return out


def extract_scalar(ws, x_col: str) -> list[dict]:
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    points = []
    for r in range(2, ws.max_row + 1):
        rd = _row_dict(ws, header, r)
        metrics = _clean_metrics(rd)
        if metrics is None:
            continue
        x = rd.get(x_col)
        if x is None:
            continue
        points.append({"x": float(x), **metrics})
    return points


def extract_pair(ws) -> list[dict]:
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    points = []
    for r in range(2, ws.max_row + 1):
        rd = _row_dict(ws, header, r)
        metrics = _clean_metrics(rd)
        if metrics is None:
            continue
        try:
            x_val = float(rd["x_value"])
            y_val = float(rd["y_value"])
        except (TypeError, ValueError, KeyError):
            continue
        points.append({
            "x_name": rd["x_name"],
            "x_value": x_val,
            "x_unit": rd["x_unit"],
            "y_name": rd["y_name"],
            "y_value": y_val,
            "y_unit": rd["y_unit"],
            **metrics,
        })
    return points


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sweeps = {}
    for sheet, cfg in SCALAR_SWEEP_SHEETS.items():
        pts = extract_scalar(wb[sheet], cfg["x_col"])
        sweeps[sheet] = {
            "x_name": cfg["x_col"],
            "x_unit": cfg["x_unit"],
            "n_points": len(pts),
            "points": pts,
        }
    pairs = {}
    for sheet in PAIR_SHEETS:
        pts = extract_pair(wb[sheet])
        pairs[sheet] = {"n_points": len(pts), "points": pts}

    payload = {
        "source_xlsx": "docs/superpowers/references/scaps_1r_parameters.xlsx",
        "source_pdf": "docs/superpowers/references/scaps_1d_simulation_report.pdf",
        "extracted_at": "2026-05-28",
        "base_model": BASE_MODEL,
        "sweeps": sweeps,
        "pairs": pairs,
    }
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, indent=2, sort_keys=False))

    print(f"Wrote {JSON_OUT.relative_to(REPO_ROOT)}")
    print(f"Base: Voc={BASE_MODEL['Voc_V']} V  Jsc={BASE_MODEL['Jsc_mA_cm2']} mA/cm2  "
          f"FF={BASE_MODEL['FF_percent']}%  PCE={BASE_MODEL['PCE_percent']}%")
    for name, blk in sweeps.items():
        print(f"  sweep {name:14s}: {blk['n_points']:3d} points")
    for name, blk in pairs.items():
        print(f"  pair  {name:24s}: {blk['n_points']:3d} points")


if __name__ == "__main__":
    main()
