"""Phase E9 — regenerate SCAPS validation overlay figures with current models.

For every single-variable sweep in the partner xlsx (1R-Parameters.xlsx), runs
SolarLab on scaps_mirror_v2.yaml with the CURRENT default models (E8 projection
optional via env; NOGEN clamp is default-on) and overlays SolarLab vs SCAPS for
V_oc / J_sc / FF / PCE. Writes one PNG per sweep.

Run from perovskite-sim/:
    python scripts/scaps_validation_figures.py --out ../docs/figures/scaps_validation
"""
from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
# Arial everywhere; mathtext in the regular (Arial) font so V$_{oc}$-style
# subscripts/superscripts render upright in Arial rather than italic Computer
# Modern. Falls back to DejaVu Sans if Arial is unavailable.
plt.rcParams["font.family"] = ["Arial", "DejaVu Sans"]
plt.rcParams["mathtext.default"] = "regular"  # math uses font.family (Arial)
plt.rcParams["axes.unicode_minus"] = True
import numpy as np
import openpyxl

from perovskite_sim.experiments.jv_sweep import run_jv_sweep
from perovskite_sim.scaps_compat import load_scaps_yaml
from perovskite_sim.sweeps.device_parameter_sweep import SweepPoint, apply_sweep_point

REPO = Path(__file__).resolve().parents[1]
CFG = REPO / "configs" / "scaps_mirror_v2.yaml"
XLSX = REPO.parent / "docs" / "superpowers" / "references" / "scaps_1r_parameters.xlsx"
JV = dict(N_grid=30, n_points=24, v_rate=5.0, V_max=1.6)

# sheet -> (updates_fn, x-axis label [mathtext], log-x, pretty title). Subscripts
# via $..$ mathtext so they render in Arial (mathtext.default=regular).
SHEETS = {
    "CHI_ETL":    (lambda v: {"etl_delta_ec_eV": v}, r"$\Delta E_C$ (eV)", False, "ETL/PVK conduction-band offset"),
    "Nd_ETL":     (lambda v: {"etl_doping_cm3": v}, r"$N_{D,ETL}$ (cm$^{-3}$)", True, "ETL donor doping"),
    "Nt_C_PVK":   (lambda v: {"absorber_defect_density_cm3": v}, r"$N_t$ PVK-CB (cm$^{-3}$)", True, "Perovskite-CB bulk defect density"),
    "Nt_V_PVK":   (lambda v: {"absorber_defect_density_cm3": v}, r"$N_t$ PVK-VB (cm$^{-3}$)", True, "Perovskite-VB bulk defect density"),
    "Nt_HTL PVK": (lambda v: {"interface_defect_N_t_cm2": v, "interface_defect_target": "htl/pvk"}, r"$N_t$ HTL/PVK (cm$^{-2}$)", True, "HTL/PVK interface defect density"),
    "Nt_PVK ETL": (lambda v: {"interface_defect_N_t_cm2": v, "interface_defect_target": "pvk/etl"}, r"$N_t$ PVK/ETL (cm$^{-2}$)", True, "PVK/ETL interface defect density"),
    "Et_C_PVK":   (lambda v: {"absorber_defect_depth_eV": v, "trap_depth_reference": "below_cb"}, r"$E_t$ PVK-CB (eV)", False, "Perovskite-CB bulk defect level"),
    "Et_V_PVK":   (lambda v: {"absorber_defect_depth_eV": v, "trap_depth_reference": "above_vb"}, r"$E_t$ PVK-VB (eV)", False, "Perovskite-VB bulk defect level"),
    "Et_HTL PVK": (lambda v: {"interface_defect_E_t_eV": v, "interface_defect_target": "htl/pvk"}, r"$E_t$ HTL/PVK (eV)", False, "HTL/PVK interface defect level"),
    "Et_PVK ETL": (lambda v: {"interface_defect_E_t_eV": v, "interface_defect_target": "pvk/etl"}, r"$E_t$ PVK/ETL (eV)", False, "PVK/ETL interface defect level"),
}


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        if d.get("scan_value") is None or d.get("Voc_V") is None:
            continue
        out.append(dict(x=float(d["scan_value"]), Voc=float(d["Voc_V"]),
                        Jsc=float(d["Jsc_mA_cm2"]), FF=float(d["FF_percent"]),
                        PCE=float(d["PCE_percent"])))
    return out


def run_sl(sheet, fn, xs):
    base = load_scaps_yaml(CFG)
    sl = {"x": [], "Voc": [], "Jsc": [], "FF": [], "PCE": []}
    for x in xs:
        sp = SweepPoint("p", sheet, f"{x:.3e}", fn(x))
        try:
            m = run_jv_sweep(apply_sweep_point(base, sp), **JV).metrics_fwd
            if not m.voc_bracketed:
                continue
            sl["x"].append(x); sl["Voc"].append(m.V_oc)
            sl["Jsc"].append(m.J_sc / 10.0)  # A/m^2 -> mA/cm^2
            sl["FF"].append(m.FF * 100.0); sl["PCE"].append(m.PCE * 100.0)
        except Exception:
            continue
    return sl


def plot_sheet(sheet, fn, xlabel, logx, title, ref, out):
    xs = [p["x"] for p in ref]
    sl = run_sl(sheet, fn, xs)
    fig, axes = plt.subplots(2, 2, figsize=(9, 6))
    metrics = [("Voc", r"$V_{oc}$ (V)"), ("Jsc", r"$J_{sc}$ (mA/cm$^2$)"),
               ("FF", "FF (%)"), ("PCE", "PCE (%)")]
    for ax, (key, ylabel) in zip(axes.ravel(), metrics):
        sc_x = [p["x"] for p in ref]; sc_y = [p[key] for p in ref]
        ax.plot(sc_x, sc_y, "k--o", ms=4, label="SCAPS", lw=1.3)
        if sl["x"]:
            ax.plot(sl["x"], sl[key], "C0-s", ms=4, label="SolarLab", lw=1.3)
        if logx:
            ax.set_xscale("log")
        ax.set_xlabel(xlabel); ax.set_ylabel(ylabel)
        ax.grid(alpha=0.3); ax.legend(fontsize=8)
    fig.suptitle(f"{title}  —  SolarLab (current models) vs SCAPS", fontsize=11)
    fig.tight_layout()
    slug = sheet.replace(" ", "_").replace("/", "_")
    p = out / f"sweep_{slug}.png"
    fig.savefig(p, dpi=110); plt.close(fig)
    print(f"  wrote {p.name}  (SL {len(sl['x'])}/{len(xs)} bracketed)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(REPO.parent / "docs" / "figures" / "scaps_validation"))
    ap.add_argument("--sheets", nargs="+", default=list(SHEETS))
    args = ap.parse_args()
    out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    for sheet in args.sheets:
        if sheet not in SHEETS or sheet not in wb.sheetnames:
            continue
        fn, xlabel, logx, title = SHEETS[sheet]
        print(f"=== {sheet} ===")
        plot_sheet(sheet, fn, xlabel, logx, title, read_sheet(wb[sheet]), out)


if __name__ == "__main__":
    main()
