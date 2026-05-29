"""Phase E9 — absolute + trend scorecard vs the partner xlsx (1R-Parameters).

Grades SolarLab against docs/superpowers/references/scaps_1r_parameters.xlsx on
BOTH absolute closeness (per-point V_oc / J_sc / FF / PCE deltas) and trend
fidelity (sweep direction + range), for the base point and every
single-variable sweep. Reads the xlsx directly (the authoritative ground
truth) rather than the derived JSON.

Env flags pass straight through to the solver so any fix combination can be
graded:  SOLARLAB_IFACE_PROJ, SOLARLAB_IFACE_EQREF, etc.

Run from perovskite-sim/:
    python scripts/scaps_absolute_scorecard.py
    SOLARLAB_IFACE_PROJ=1 python scripts/scaps_absolute_scorecard.py --sheets CHI_ETL Nd_ETL
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

from perovskite_sim.experiments.jv_sweep import run_jv_sweep
from perovskite_sim.scaps_compat import load_scaps_yaml
from perovskite_sim.sweeps.device_parameter_sweep import SweepPoint, apply_sweep_point

REPO = Path(__file__).resolve().parents[1]
CFG = REPO / "configs" / "scaps_mirror_v2.yaml"
XLSX = REPO.parent / "docs" / "superpowers" / "references" / "scaps_1r_parameters.xlsx"

# sheet -> (axis-updates builder). Mirrors run_scaps_full_regression mapping.
SHEETS = {
    "CHI_ETL": lambda v: {"etl_delta_ec_eV": v},
    "Nd_ETL": lambda v: {"etl_doping_cm3": v},
    "Nt_C_PVK": lambda v: {"absorber_defect_density_cm3": v},
    "Nt_V_PVK": lambda v: {"absorber_defect_density_cm3": v},
    "Nt_HTL PVK": lambda v: {"interface_defect_N_t_cm2": v, "interface_defect_target": "htl/pvk"},
    "Nt_PVK ETL": lambda v: {"interface_defect_N_t_cm2": v, "interface_defect_target": "pvk/etl"},
    "Et_C_PVK": lambda v: {"absorber_defect_depth_eV": v, "trap_depth_reference": "below_cb"},
    "Et_V_PVK": lambda v: {"absorber_defect_depth_eV": v, "trap_depth_reference": "above_vb"},
    "Et_HTL PVK": lambda v: {"interface_defect_E_t_eV": v, "interface_defect_target": "htl/pvk"},
    "Et_PVK ETL": lambda v: {"interface_defect_E_t_eV": v, "interface_defect_target": "pvk/etl"},
}

JV = dict(N_grid=30, n_points=20, v_rate=5.0, V_max=1.6)


def read_xlsx_sheet(ws):
    """Return list of dicts: x (scan_value), Voc, Jsc_A_m2, FF, PCE."""
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        if d.get("scan_value") is None or d.get("Voc_V") is None:
            continue
        out.append({
            "x": float(d["scan_value"]),
            "Voc": float(d["Voc_V"]),
            "Jsc": float(d["Jsc_mA_cm2"]) * 10.0,  # mA/cm^2 -> A/m^2
            "FF": float(d["FF_percent"]) / 100.0,
            "PCE": float(d["PCE_percent"]) / 100.0,
        })
    return out


def grade(sheet, fn, ref):
    base = load_scaps_yaml(CFG)
    sl = []
    for pt in ref:
        sp = SweepPoint("p", sheet, f"{pt['x']:.3e}", fn(pt["x"]))
        try:
            m = run_jv_sweep(apply_sweep_point(base, sp), **JV).metrics_fwd
            sl.append((pt, m))
        except Exception as e:
            sl.append((pt, None))
    brk = [(pt, m) for pt, m in sl if m is not None and m.voc_bracketed]
    if not brk:
        return None
    # absolute: median |Δ| over bracketed points
    voc_abs = sorted(abs(m.V_oc - pt["Voc"]) * 1000 for pt, m in brk)
    jsc_abs = sorted(abs(m.J_sc - pt["Jsc"]) for pt, m in brk)
    med = lambda a: a[len(a) // 2]
    # trend: range + net direction
    sl_voc = [m.V_oc for _, m in brk]
    sc_voc = [pt["Voc"] for pt, _ in brk]
    sl_rng = (max(sl_voc) - min(sl_voc)) * 1000
    sc_rng = (max(sc_voc) - min(sc_voc)) * 1000
    closure = 100 * sl_rng / sc_rng if sc_rng > 1e-6 else float("nan")
    dir_ok = (sl_voc[-1] - sl_voc[0] >= 0) == (sc_voc[-1] - sc_voc[0] >= 0)
    return dict(n=len(brk), voc_abs_med=med(voc_abs), jsc_abs_med=med(jsc_abs),
                sc_rng=sc_rng, sl_rng=sl_rng, closure=closure, dir_ok=dir_ok)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheets", nargs="+", default=list(SHEETS))
    args = ap.parse_args()
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # base point
    base_m = run_jv_sweep(load_scaps_yaml(CFG), **JV).metrics_fwd
    bsheet = read_xlsx_sheet(wb["Nt_HTL PVK"])  # base = the Nt=1e12 row, ~all sheets agree
    sc_base = next(p for p in bsheet if abs(p["x"] - 1e12) < 1)
    print(f"BASE: V_oc {base_m.V_oc:.4f}/{sc_base['Voc']:.4f} (Δ{(base_m.V_oc-sc_base['Voc'])*1000:+.0f}mV)  "
          f"J_sc {base_m.J_sc:.0f}/{sc_base['Jsc']:.0f} (Δ{base_m.J_sc-sc_base['Jsc']:+.0f})  "
          f"FF {base_m.FF:.3f}/{sc_base['FF']:.3f}  PCE {base_m.PCE*100:.2f}/{sc_base['PCE']*100:.2f}%")
    print()
    print(f"{'sheet':14s} {'n':>2s} {'VocΔmed':>8s} {'JscΔmed':>8s} {'closure':>8s} {'dir':>4s}")
    for sheet in args.sheets:
        if sheet not in SHEETS or sheet not in wb.sheetnames:
            continue
        st = grade(sheet, SHEETS[sheet], read_xlsx_sheet(wb[sheet]))
        if st is None:
            print(f"{sheet:14s}  insufficient")
            continue
        print(f"{sheet:14s} {st['n']:>2d} {st['voc_abs_med']:>7.0f}m {st['jsc_abs_med']:>7.0f} "
              f"{st['closure']:>7.0f}% {'ok' if st['dir_ok'] else 'X':>4s}")


if __name__ == "__main__":
    main()
